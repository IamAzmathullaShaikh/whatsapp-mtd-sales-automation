"""
MTD file schema detection: turn *any* alcohol-beverage MTD dump (or party
master) into the app's canonical schema without hardcoding sheet names, header
rows, or column spellings.

Detection is fuzzy on purpose: it scans the first few sheets and header rows,
matches column headers against synonym sets for each *role* (depot, syndicate,
vendor, total, party, phone, ...), and returns a proposed mapping the user can
confirm or edit in the GUI.

Pure and dependency-light (openpyxl only), so it is unit-testable by writing
throwaway workbooks in a tmp dir.
"""

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Role synonyms — add real-world spellings here as new companies onboard.
# Headers are normalized (lowercase, punctuation stripped) before matching.
# ---------------------------------------------------------------------------
SALES_ROLES = {
    "depot":     ["name of depot", "depot", "depot name", "godown", "warehouse", "branch", "area"],
    "syndicate": ["syndicate name", "syndicate", "group name", "group", "outlet group", "party group", "cluster"],
    "vendor":    ["vendor name", "vendor", "outlet", "customer", "customer name", "party", "distributor", "account", "shop"],
    "total":     ["total", "total qty", "total cases", "total volume", "grand total", "qty", "cases", "volume", "units", "quantity"],
}

PARTY_ROLES = {
    "party":        ["party", "party name", "account", "account name", "customer", "customer name", "distributor", "name", "outlet"],
    "phone":        ["phone", "mobile", "mobile no", "mobile number", "phone number", "contact", "contact no", "whatsapp", "whatsapp no", "number"],
    "send":         ["send", "report", "send report", "receive report", "opt in", "whatsapp report", "report status"],
    "priority":     ["priority", "tier", "grade", "slab", "class", "segment"],
    "total_target": ["total target", "monthly target", "target total", "target", "mtd target"],
}

HEADER_SCAN_ROWS = 12  # how deep to hunt for the header row in each sheet
MAX_SHEETS = 5        # don't scan the whole workbook on big files


def normalize_header(value):
    """'Name Of Depot' -> 'nameofdepot' (lowercase, no spaces/punctuation)."""
    if value is None:
        return ""
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def _match_headers(values, roles):
    """Return {role: actual_header} for every role whose synonyms matched."""
    found = {}
    for role, synonyms in roles.items():
        for syn in synonyms:
            needle = normalize_header(syn)
            for header in values:
                hay = normalize_header(header)
                if hay and (needle in hay or (len(hay) >= 4 and hay in needle)):
                    found[role] = header
                    break
            if role in found:
                break
    return found


def sheet_names(file_path):
    """Workbook sheet names (for the GUI dropdown)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def header_row_values(file_path, sheet, row):
    """Non-empty cell values at (sheet, row) — the candidate column headers."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        for values in ws.iter_rows(min_row=row + 1, max_row=row + 1, values_only=True):
            return [str(v).strip() for v in values if v is not None and str(v).strip()]
        return []
    finally:
        wb.close()


def _detect(file_path, roles, min_roles):
    """Shared detector: best (sheet, header_row, {role: header}) across the workbook."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        best = None  # (score, row_idx, sheet, found)
        for sheet in wb.worksheets[:MAX_SHEETS]:
            for row_idx, row in enumerate(sheet.iter_rows(
                    min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True)):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if not values:
                    continue
                found = _match_headers(values, roles)
                score = len(found)
                if best is None or score > best[0]:
                    best = (score, row_idx, sheet.title, found)
        if best is None:
            return None
        score, row_idx, sheet_name, found = best
        if score < min_roles:
            return None
        return {
            "sheet": sheet_name,
            "header_row": row_idx,
            "columns": found,
            "confidence": round(score / len(roles), 2),
        }
    finally:
        wb.close()


def detect_sales_schema(file_path, roles=None):
    """Detect the sales (MTD) dump schema — needs >= 3 of depot/syndicate/vendor/total."""
    return _detect(file_path, roles or SALES_ROLES, min_roles=3)


def detect_party_schema(file_path):
    """Detect the party master schema — needs >= 2 of party/phone/send/priority/target."""
    return _detect(file_path, PARTY_ROLES, min_roles=2)


def apply_mapping(columns, actual_columns):
    """
    Given {role: actual_header} plus the actual column list, return the rename
    map {canonical_name: actual_header} for roles whose header is present.
    Used by pipeline.load_dataframes to normalize any file into the canonical
    schema the rest of the pipeline expects.
    """
    from config import (COL_DEPOT, COL_SYNDICATE, COL_VENDOR, COL_TOTAL,
                        COL_PARTY, COL_PHONE, COL_SEND, COL_PRIORITY,
                        TOTAL_TARGET_COL)
    canonical = {
        "depot": COL_DEPOT, "syndicate": COL_SYNDICATE, "vendor": COL_VENDOR,
        "total": COL_TOTAL, "party": COL_PARTY, "phone": COL_PHONE,
        "send": COL_SEND, "priority": COL_PRIORITY, "total_target": TOTAL_TARGET_COL,
    }
    rename = {}
    for role, actual in (columns or {}).items():
        if actual in actual_columns and canonical.get(role):
            rename[canonical[role]] = actual
    return rename
